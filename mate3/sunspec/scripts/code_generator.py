import re
from collections import OrderedDict
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from loguru import logger
from mate3.sunspec.fields import Field, Mode
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
PATH = ROOT / "doc" / "OutBack.Power.SunSpec.Map.xlsx"
REGISTERS_SHEET = "SunSpecMap"
BITFIELDS_SHEET = "Bitfields"
MODELS_MODULE = ROOT / "models.py"
VALUES_MODULE = ROOT / "values.py"
MODELS_COLUMNS = OrderedDict(
    [
        ("did", {"column_names": ("DID",), "type": int}),
        ("start", {"column_names": ("Start",), "type": int}),
        ("end", {"column_names": ("End",), "type": int}),
        ("size", {"column_names": ("Size",), "type": int}),
        ("mode", {"column_names": ("R/W",), "type": str}),
        ("name", {"column_names": ("Name", "Field Name"), "type": str}),
        ("type", {"column_names": ("Type",), "type": str}),
        ("units", {"column_names": ("Units",), "type": str}),
        ("scale_factor", {"column_names": ("Scale Factor",), "type": str}),
        ("contents", {"column_names": ("Contents",), "type": str}),
        ("description", {"column_names": ("Description",), "type": str}),
    ]
)
BITFIELDS_COLUMNS = ("did", "bitfield_name", "name", "mask", "value", "description")

WARNING = f"This file is auto generated, do not edit. The generation code can be found in {Path(__file__).name}"


@dataclass
class ModelRow:

    did: int
    start: int
    end: int
    size: int
    mode: str
    name: str
    type: str
    units: str
    scale_factor: str
    contents: str
    description: str

    # set dynamically:
    python_name: str = None


@dataclass(frozen=False)
class BitfieldRow:

    did: int
    bitfield_name: str
    name: str
    mask: int
    value: int
    description: str

    # set dynamically:
    python_name: str = None


def python_name_from_field(name):
    name = name.strip()
    # first, replace kW/kWh (and variants of caps) with kw/kwh to avoid splitting:
    name = re.sub(r"(^|_)kwh(_|$)", r"\1kwh\2", name, flags=re.I)
    name = re.sub(r"(^|_)kw(_|$)", r"\1kw\2", name, flags=re.I)
    # now do naive snake case underscore insert - <lower><UPPER> -> <lower>_<upper>. Why naive? That's all we need.
    name = re.sub(r"([a-z])([A-Z])", r"\1_\2", name).lower()
    # remove ick:
    name = re.sub(r"[^a-zA-Z0-9_]", r"", name)
    # manual renamings:
    name = re.sub(r"batt(_|$)", r"battery\1", name)
    name = re.sub(r"_sf$", r"_scale_factor", name)
    name = re.sub(r"_temp(_|$)", r"_temperature\1", name)
    return name


def find_common_prefixes(rows):
    """
    Find any common prefixes for the registry names given in the provided lines, after splitting on _. E.g.

    PREFIX_a
    PREFIX_b_c
    ...

    Would return {'PREFIX'}. While

    PREFIX_PREFIX2_a
    PREFIX_PREFIX2_b_c
    ...

    Would return {'PREFIX_PREFIX2'}. We do this by finding the common ancestor that all have. If there is not at
    root level, then we just return the prefixes at root level, i.e.

    PREFIX1_a
    PREFIX2_b_c
    ...

    Would return {'PREFIX1', 'PREFIX2'}
    """

    if len(rows) <= 1:
        return set()

    splats = [row.name.split("_") for row in rows]
    previous_common_ancestor = None
    for i in range(1, max(len(s) for s in splats)):
        prefixes = set(["_".join(s[:i]) + "_" for s in splats])
        if len(prefixes) > 1:
            return prefixes if previous_common_ancestor is None else previous_common_ancestor
        previous_common_ancestor = prefixes
    else:
        raise RuntimeError("Couldn't find common prefixes!")


def strip_prefix(name, common_prefixes):
    """snake_case + remove common prefix at start"""
    for prefix in common_prefixes:
        if name.startswith(prefix):
            return name.replace(prefix, "", 1)
    return name


def clean_rows(rows):
    common_prefixes = find_common_prefixes(rows)
    for row in rows:
        row.python_name = python_name_from_field(strip_prefix(row.name, common_prefixes))


class ModelTable:
    def __init__(self, *, name, rows):
        self.name = name
        if len({row.did for row in rows}) > 1:
            raise ValueError("All rows must have the same DID!")
        self.did = rows[0].did
        self.rows = rows
        clean_rows(self.rows)

    @property
    def python_name(self):
        name = re.sub(r"[\s\-]", "", self.name)
        name = re.sub(r"(block|model)$", "", name, flags=re.I)
        return re.sub(r"(block|model)$", "", name, flags=re.I)

    def generate_definition(self, bitfields):
        class_name = self.python_name
        code = f"""class {class_name}Model(Model):
    
    def __init__(self):
"""

        # Process rows:
        processed = []
        for row in self.rows:
            processed.append(self._generate_field(row=row, class_name=class_name, bitfields=bitfields))

        # Sort them so scale factors come at the end:
        processed = sorted(processed, key=lambda x: x[2])
        for name, defn, scale_factor in processed:
            code += defn
        return code

    def _generate_enumerated_field(self, row):
        # ok, generate always seems to following the form 1=..., 2=..., etc. or ...=1, ...=2, So let's look for that
        starts_number = list(re.finditer(r"(\-?[0-9]+)\s?=\s?(.*?)(?=\-?[0-9]+\s?=|$)", row.description))
        ends_number = list(re.finditer(r"(.*?)\s?=\s?(\-?[0-9]+)", row.description))
        options = []
        use_start = True
        if starts_number:
            if ends_number:
                n_start = len(starts_number)
                n_end = len(ends_number)
                if n_start == n_end:
                    # ok, confusing ... must be something like 0 = 1, 1 = 2, etc.
                    pass
                else:
                    # use the one with the most:
                    use_start = n_start > n_end
                    logger.warning(
                        f"field {row.name} is enumerated but the description matches both number-first and "
                        f"number-last enumeration. Going with number-{'first' if use_start else 'last'} as there are "
                        f"{n_start} number-first matches and {n_end} number-last matches.\n\tNB desc: {row.description}"
                    )
        elif ends_number:
            use_start = False
        else:
            raise ValueError("Couldn't match description to enumeration ...")

        if use_start:
            for m in starts_number:
                g = m.groups()
                k = int(g[0])
                v = g[1].strip().strip(",;").strip()
                options.append((v, k))
        else:
            for m in ends_number:
                g = m.groups()
                k = int(g[1])
                v = g[0].strip().strip(",")
                options.append((v, k))

        # check all values are unique:
        if len(set(i[1] for i in options)) < len(options):
            raise ValueError("Duplicate numeric values in enumerable!")

        # get rid of duplicates:
        unique_options = {}
        p = re.compile(r"[^a-z0-9]+", flags=re.I)
        for v, k in options:
            v_original = p.sub("_", v).upper().strip("_")
            if v in unique_options:
                v = f"{v_original}_{k}"
                logger.warning(f"{v_original}={k} renamed as {v}={k} as {v} already exists")
            unique_options[v] = k

        enum = f'Enum("{row.python_name}", {str(list(sorted(unique_options.items())))})'
        yield f"enum={enum}"

    def _generate_field(self, row, class_name, bitfields):
        """Generate a single Field definition for a table Model class"""

        name = row.python_name
        if name in ("sun_spec_did", "sun_spec_length"):
            name = name.replace("sun_spec_", "")
        field_type = None
        field_args = [f'"{name}"', f"{row.start}", f"{row.size}", f"{Mode(row.mode.lower().replace('/', ''))}"]
        has_scale_factor = False
        if row.type in ("uint16", "int16", "uint32", "int32"):
            int_field = True
            units = row.units.lower().strip() if row.units is not None else None
            if units == "bitfield":
                int_field = False
                # if we don't know about that bitfield, skip it - generally these are ones kept for future use
                if row.name not in bitfields:
                    logger.warning(f"skipping {row.name} as unknown bitfield. Row: {row}")
                    return "", "", has_scale_factor
                field_type = f"Bit{16 * row.size}"
                field_args.append(f"flags={bitfields[row.name]}")
            elif units == "enumerated" or (
                row.description is not None
                and re.match(r"^\s*0\s*=\s*Disabled\s*[,;]\s*1\s*=\s*Enabled\s*$", row.description)
            ):
                try:
                    field_args += list(self._generate_enumerated_field(row))
                    field_type = f"Enum{row.type.title()}"
                    int_field = False
                except ValueError as e:
                    logger.warning(
                        (
                            f"field {row.name} is enumerated but the description isn't enumerable so "
                            f"treating as int: \n\tDesc: '{row.description}'\n\tError: {e}"
                        )
                    )
            elif units == "address":
                int_field = False
                field_type = "Address"

            if int_field:
                if row.units:
                    field_args.append(f'units="{row.units}"')
                field_type = row.type.title()
                if row.scale_factor is not None:
                    scale_factors = [r for r in self.rows if r.name == row.scale_factor]
                    if len(scale_factors) != 1:
                        raise RuntimeError("Couldn't get scale factors!")
                    field_args.append(f"scale_factor=self.{scale_factors[0].python_name}")
                    field_type = f"Float{field_type}"
                    has_scale_factor = True
        elif row.type.startswith("string"):
            field_type = "String"
        else:
            raise ValueError(f"Don't know what to do with type {row.type}")

        # Add description at the end
        if row.description:
            field_args.append(f'description="{row.description}"')

        return (
            name,
            f"        self.{name}: {field_type}Field = {field_type}Field({', '.join(field_args)})\n",
            has_scale_factor,
        )


class BitfieldTable:
    def __init__(self, rows):
        did = rows[0].did
        name = rows[0].bitfield_name
        for row in rows:
            if row.did != did or row.bitfield_name != name:
                raise RuntimeError("All rows should have the same name and did")
        self.name = name
        self.rows = self._sanitise_rows(rows)
        clean_rows(self.rows)

    def _sanitise_rows(self, rows):
        """
        Sometimes the 'off' value is specified in the table too e.g.
            Virtual Name            Mask    Value   Description
            OB_Inverter_AC_Input    0x0004  x0004   Inverter AC Input Use
            OB_Inverter_AC_Input    0x0004  0x0000  Inverter AC Input Drop
        In this case, we only keep the 'on' one (where value is !0) and then update the description of value + mask.
        """
        on_by_name = {}
        off_by_name = {}
        for row in rows:
            if row.value is None or int(row.value, base=16) != 0:
                if row.name in on_by_name:
                    raise RuntimeError("multiple on rows!")
                on_by_name[row.name] = row
            else:
                if row.name in off_by_name:
                    raise RuntimeError("multiple off rows!")
                off_by_name[row.name] = row

        # update the ons with the offs:
        for row in off_by_name.values():
            logger.warning(
                (
                    f"{row.name} has an on value ({on_by_name[row.name].description}), and an off ({row.description}). "
                    "Assuming these are opposites, so ignoring the off value so we have a proper bitfield."
                )
            )
            on_by_name[row.name].description += f" (Unset means '{row.description}')"

        return sorted(list(on_by_name.values()), key=lambda x: x.mask)

    @property
    def python_name(self):
        name = self.name.replace("_", "")
        name = re.sub(r"flags$", "", name, flags=re.I)
        return name + "Flags"

    def generate_definition(self):
        class_name = self.python_name
        code = f"@unique\nclass {class_name}(DescribedIntFlag):\n"
        for row in self.rows:
            mask = int(row.mask, base=16)
            code += f'    {row.python_name} = {mask}, "{row.description}"\n'
        return code


# function to ensure string values are ascii, and N/A -> null
def sanitise(value):
    if isinstance(value, str):
        value = value.encode("ascii", errors="backslashreplace").decode("ascii").strip()
        return None if value == "N/A" else value
    return value


def read_model_table(row_iter):

    # Start looking though rows until we find the start of a table - that is, where DID column == 'DID'. If we don't see
    # this for 100 rows (as there are some 'empty' error tables in between) then we assume there are no more tables.
    more_tables = True
    row = None
    previous_row = None
    try:
        for _ in range(100):
            row = next(row_iter)
            named_values = {k: row[idx].value for idx, k in enumerate(MODELS_COLUMNS)}
            if named_values["did"] == "DID":
                break
            previous_row = row
        else:
            # Ok, more than 100 blank lines - there aren't any tables after this:
            more_tables = False
    except StopIteration:
        # Or the iterator finished (as it was clever enough to know there was nothing else)
        more_tables = False

    if not more_tables:
        return None, None

    title_row = previous_row
    table_name = re.sub("^Table [0-9]+", "", title_row[1].value).strip()

    # Cool, next row is header - check column names:
    header_row = row
    for cidx, (name, (_, col)) in enumerate(zip(header_row, MODELS_COLUMNS.items())):
        name = name.value
        expected_names = col["column_names"]
        if name not in expected_names:
            raise ValueError(f"Expected table header for column {cidx} to be in {expected_names} but was '{name}'")

    # The rest of the rows, until we hit blank DID, are fields:
    model_rows = []
    for row in row_iter:
        named_values = {k: sanitise(row[idx].value) for idx, k in enumerate(MODELS_COLUMNS)}
        # if this is the first row, and it's blank, skip it - sometimes there's just a completely empty row after the
        # header:
        if not model_rows and all(v is None for v in named_values.values()):
            continue
        # if DID cell is blank, then we've finished parsing table:
        if named_values["did"] is None:
            return table_name, model_rows
        # type it:
        named_values = {k: v if v is None else MODELS_COLUMNS[k]["type"](v) for k, v in named_values.items()}
        model_rows.append(ModelRow(**named_values))
    # in case we finish iter:
    return table_name, model_rows


def read_bitfields(wb):
    did_names = {}
    ws = wb[BITFIELDS_SHEET]
    current_did = None
    current_name = None
    for row in ws.iter_rows(max_col=len(BITFIELDS_COLUMNS) + 1):
        values = {k: sanitise(v.value) for k, v in zip(BITFIELDS_COLUMNS, row)}
        did = values["did"]
        name = values["bitfield_name"]
        if isinstance(did, int):
            if did != current_did or name != current_name:
                if (did, name) in did_names:
                    raise RuntimeError("Uh-oh, repeat DID + Name table!")
                did_names[(did, name)] = []
            did_names[(did, name)].append(BitfieldRow(**values))
            current_did = did
            current_name = name
        else:
            current_did = None
            current_name = None
    tables = {}
    for rows in did_names.values():
        table = BitfieldTable(rows=rows)
        # last I_Event_1 ones seem to be repeats, just for different DID's - let's check they are the same
        if table.name in tables:
            a = [BitfieldRow(did=1, **{k: v for k, v in r.__dict__.items() if k != "did"}) for r in table.rows]
            b = [
                BitfieldRow(did=1, **{k: v for k, v in r.__dict__.items() if k != "did"})
                for r in tables[table.name].rows
            ]
            if a != b:
                raise RuntimeError(f"Duplicate table of name {table.name} but different rows.")
        else:
            tables[table.name] = table
    return list(tables.values())


def read_models(wb):
    tables = {}
    ws = wb[REGISTERS_SHEET]
    row_iter = ws.iter_rows(max_col=len(MODELS_COLUMNS) + 1)
    while True:
        name, rows = read_model_table(row_iter)
        if rows is None:
            break
        if name == "UPS Inverters":
            # This table is weird. Ignore it.
            break
        table = ModelTable(name=name, rows=rows)
        if table.name in tables:
            raise ValueError(f"Two tables named {table.name}")
        tables[table.name] = table
    return list(tables.values())


def main():

    wb = load_workbook(PATH)

    code = f'''"""{WARNING}"""


from enum import Enum, IntFlag, unique
from mate3.sunspec.fields import (
    Mode,
    StringField,
    Int16Field,
    Uint16Field,
    Uint32Field,
    FloatInt16Field,
    FloatUint16Field,
    FloatUint32Field,
    EnumUint16Field,
    EnumInt16Field,
    Bit16Field,
    Bit32Field,
    DescribedIntFlag,
    AddressField
)
from mate3.sunspec.model_base import Model


'''
    # bitfields first:
    bitfield_tables = read_bitfields(wb)
    bitfield_tables = sorted(bitfield_tables, key=lambda x: x.python_name)
    for table in bitfield_tables:
        code += table.generate_definition()
        code += "\n\n"

    code += f"""class SunSpecHeaderModel(Model):
    def __init__(self):
        self.did: Uint32Field = Uint32Field("did", 1, 2, Mode.R)
        self.model_id: Uint16Field = Uint16Field("model_id", 3, 1, Mode.R)
        self.length: Uint16Field = Uint16Field("length", 4, 1, Mode.R)
        self.manufacturer: StringField = StringField("manufacturer", 5, 16, Mode.R)
        self.model: StringField = StringField("model", 21, 16, Mode.R)
        self.options: StringField = StringField("options", 37, 8, Mode.R)
        self.version: StringField = StringField("version", 45, 8, Mode.R)
        self.serial_number: StringField = StringField("serial_number", 53, 16, Mode.R)


class SunSpecEndModel(Model):
    def __init__(self):
        self.did: Uint16Field = Uint16Field("did", 1, 1, Mode.R, description="Should be {0xFFFF}")
        self.length: Uint16Field = Uint16Field("length", 2, 1, Mode.R, description="Should be 0")


"""
    model_tables = read_models(wb)
    model_tables = sorted(model_tables, key=lambda x: x.did)
    bitfield_lookup = {t.name: t.python_name for t in bitfield_tables}
    for table in model_tables:
        code += table.generate_definition(bitfield_lookup)
        code += "\n\n"

    code += "MODEL_DEVICE_IDS = {\n"
    code += f"    {0x53756e53}: SunSpecHeaderModel,\n"
    code += "    65535: SunSpecEndModel,\n"
    for table in model_tables:
        code += f"    {table.did}: {table.python_name}Model,\n"
    code += "}\n"

    with open(str(MODELS_MODULE), "w") as f:
        f.write(code)


if __name__ == "__main__":
    main()
